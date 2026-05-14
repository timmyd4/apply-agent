#!/usr/bin/env python3
import os
import re
import shutil
import subprocess
import sys
import tempfile
import yaml
from google import genai
from google.genai import types
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

GEMINI_FLASH_15  = "gemini-1.5-flash"   # free tier fallback
GEMINI_FLASH_20  = "gemini-2.0-flash"   # no free tier
GEMINI_FLASH_25  = "gemini-2.5-flash"   # best quality, 20 req/day free tier

GEMINI_MODEL = GEMINI_FLASH_15          # change this line to switch models

ROOT = Path(__file__).parent
MASTER_RESUME = ROOT / "Tim_Williams_Master_Resume.tex"
SKILLS_DIR = ROOT / "skills"
JOB_DESCRIPTION = ROOT / "job_description.md"
INSTRUCTIONS = ROOT / "instructions.md"
OUTPUT_DIR = ROOT / "output"
TRACKER = ROOT / "applications.xlsx"

HEADERS = ["Date", "Company", "Role", "Resume File", "Status", "Notes"]
HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STATUS_COLORS = {
    "Generated": "FFFFFF",
    "Applied":   "D9EAD3",
    "Interview": "FFF2CC",
    "Offer":     "B6D7A8",
    "Rejected":  "F4CCCC",
}


def update_tracker(company: str, role: str, filename: str) -> None:
    if TRACKER.exists():
        wb = load_workbook(TRACKER)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 30

    row = [
        datetime.now().strftime("%Y-%m-%d"),
        company.replace("_", " "),
        role.replace("_", " "),
        filename,
        "Generated",
        "",
    ]
    ws.append(row)

    last_row = ws.max_row
    fill = PatternFill("solid", fgColor=STATUS_COLORS["Generated"])
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=last_row, column=col).fill = fill

    wb.save(TRACKER)


def load_skills() -> str:
    skills = []
    for yaml_file in sorted(SKILLS_DIR.glob("*.yaml")):
        with open(yaml_file, encoding="utf-8") as f:
            skills.append(f"# {yaml_file.stem}\n{f.read()}")
    return "\n---\n".join(skills)


def extract_metadata(job_desc: str) -> tuple[str, str]:
    match = re.match(r"^---\n(.*?)\n---", job_desc, re.DOTALL)
    if match:
        try:
            meta = yaml.safe_load(match.group(1))
            if meta:
                company = re.sub(r"[^\w\-]", "_", str(meta.get("company", "Company")).strip())
                role = re.sub(r"[^\w\-]", "_", str(meta.get("role", "Role")).strip())
                return company, role
        except Exception:
            pass
    return "Company", "Role"


COVER_LETTER_INSTRUCTION = (
    "Write a short cover letter paragraph for Tim Williams, a software engineering student applying for an internship. "
    "Write in first person. It should sound like a real college student wrote it — simple, direct, and genuine. "
    "Say that Tim works hard, cares about doing things right, and is eager to learn as much as he can while building real software. "
    "Mention one specific thing from the job description to make it feel personal, not generic. "
    "Keep it to 3 sentences max. "
    "Do NOT use any of these words or phrases: impactful, leverage, synergy, tangible, business value, real-world setting, contribute meaningfully, "
    "clean reliable systems, operational workflows, drive value, passionate, thrilled, delighted, excited to announce, unique opportunity. "
    "Do not mirror the job description's own language back at it. Do not reference company values by name. "
    "Write like a person, not like ChatGPT wrote a cover letter. Short words, short sentences, no corporate polish. "
    "Return only the paragraph — no salutation, no closing, no extra formatting."
)

COVER_LETTER_TEMPLATE = r"""\documentclass[12pt]{{article}}
\usepackage[margin=1in]{{geometry}}
\usepackage{{times}}
\usepackage{{parskip}}
\pagestyle{{empty}}

\begin{{document}}

\textbf{{Tim Williams}} \hfill timmywilliams4665@gmail.com

\vspace{{1em}}

{date}

\vspace{{1em}}

Hiring Manager \\
{company}

\vspace{{1em}}

Dear Hiring Manager,

{paragraph}

\vspace{{1em}}

Sincerely,

\vspace{{2em}}

Tim Williams

\end{{document}}
"""


def generate_cover_letter(client, company: str, role: str, job_description: str, timestamp: str) -> Path:
    prompt = (
        f"Company: {company.replace('_', ' ')}\n"
        f"Role: {role.replace('_', ' ')}\n\n"
        f"Job Description:\n{job_description}"
    )

    print(f"Generating cover letter for {company} — {role}...")

    response = client.models.generate_content(
        model=GEMINI_MODEL,
        config=types.GenerateContentConfig(
            system_instruction=COVER_LETTER_INSTRUCTION,
            max_output_tokens=8192,
        ),
        contents=prompt,
    )

    paragraph = response.text.strip()

    tex = COVER_LETTER_TEMPLATE.format(
        date=datetime.now().strftime("%B %d, %Y"),
        company=company.replace("_", " "),
        paragraph=paragraph,
    )

    out_file = OUTPUT_DIR / f"Tim_Williams_{company}_{role}_CoverLetter_{timestamp}.tex"
    out_file.write_text(tex, encoding="utf-8")
    return out_file


TRIM_INSTRUCTION = (
    "The LaTeX resume you produced compiled to more than one page. "
    "Trim it to fit exactly one page by cutting bullets and shortening text — "
    "do NOT change any LaTeX formatting commands, packages, or spacing. "
    "Return only the complete raw LaTeX source from \\documentclass to \\end{document}."
)


def compile_and_count_pages(tex_source: str) -> int | None:
    """
    Compile tex_source with pdflatex in a temp dir and return the page count.
    Returns None if pdflatex is not installed or compilation fails.
    """
    if not shutil.which("pdflatex"):
        return None
    with tempfile.TemporaryDirectory() as tmp:
        tex_file = Path(tmp) / "resume.tex"
        tex_file.write_text(tex_source, encoding="utf-8")
        result = subprocess.run(
            ["pdflatex", "-interaction=nonstopmode", "-halt-on-error", str(tex_file)],
            capture_output=True,
            text=True,
            cwd=tmp,
        )
        # pdflatex prints: "Output written on resume.pdf (N page(s), ...)"
        match = re.search(r"Output written on .+?\((\d+) page", result.stdout)
        if match:
            return int(match.group(1))
        return None


def enforce_one_page(client, tex_source: str, prompt: str, instructions: str) -> str:
    """
    Compile tex_source; if it exceeds 1 page, ask Gemini to trim and retry up to 2 times.
    Returns the final (possibly trimmed) LaTeX source.
    """
    for attempt in range(3):
        pages = compile_and_count_pages(tex_source)
        if pages is None:
            if attempt == 0:
                print(
                    "WARNING: pdflatex not found — skipping page-count verification. "
                    "Install MiKTeX or TeX Live to enable the compile loop."
                )
            break
        if pages <= 1:
            if attempt > 0:
                print(f"Trimmed to 1 page after {attempt} retry(ies).")
            break
        print(f"Resume compiled to {pages} pages — asking Gemini to trim (attempt {attempt + 1}/2)...")
        trim_prompt = (
            f"## Current LaTeX (compiled to {pages} pages)\n```latex\n{tex_source}\n```\n\n"
            f"## Original Prompt Context\n{prompt}"
        )
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            config=types.GenerateContentConfig(
                system_instruction=instructions + "\n\n" + TRIM_INSTRUCTION,
                max_output_tokens=16384,
            ),
            contents=trim_prompt,
        )
        tex_source = response.text.strip()
        if tex_source.startswith("```"):
            lines = tex_source.splitlines()
            end = len(lines) - 1 if lines[-1].strip() == "```" else len(lines)
            tex_source = "\n".join(lines[1:end])
    else:
        print("WARNING: Still >1 page after 2 trim attempts. Check the .tex file manually.")
    return tex_source


def main():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY environment variable is not set.")
        sys.exit(1)

    for path in [MASTER_RESUME, JOB_DESCRIPTION, INSTRUCTIONS]:
        if not path.exists():
            print(f"ERROR: Missing required file: {path}")
            sys.exit(1)

    master_resume = MASTER_RESUME.read_text(encoding="utf-8")
    job_description = JOB_DESCRIPTION.read_text(encoding="utf-8").strip()
    instructions = INSTRUCTIONS.read_text(encoding="utf-8")
    skills_yaml = load_skills()

    if not job_description or job_description.endswith("-->"):
        print("ERROR: job_description.md appears empty — paste the job posting below the comment.")
        sys.exit(1)

    company, role = extract_metadata(job_description)

    client = genai.Client(api_key=api_key)

    prompt = (
        f"## Master Resume\n```latex\n{master_resume}\n```\n\n"
        f"## Skill Library\n```yaml\n{skills_yaml}\n```\n\n"
        f"## Job Description\n\n{job_description}"
    )

    print(f"Generating resume for {company} — {role}...")

    response = client.models.generate_content(
        model=GEMINI_MODEL,
        config=types.GenerateContentConfig(
            system_instruction=instructions,
            max_output_tokens=16384,
        ),
        contents=prompt,
    )

    tailored_tex = response.text.strip()

    # Strip markdown fences if the model wrapped the output anyway
    if tailored_tex.startswith("```"):
        lines = tailored_tex.splitlines()
        end = len(lines) - 1 if lines[-1].strip() == "```" else len(lines)
        tailored_tex = "\n".join(lines[1:end])

    if not tailored_tex.endswith(r"\end{document}"):
        print("WARNING: Resume output appears truncated — \\end{document} not found. Check the .tex file before committing.")

    tailored_tex = enforce_one_page(client, tailored_tex, prompt, instructions)

    OUTPUT_DIR.mkdir(exist_ok=True)
    for old in OUTPUT_DIR.glob("*.tex"):
        old.unlink()

    timestamp = datetime.now().strftime("%Y-%m-%d")
    out_file = OUTPUT_DIR / f"Tim_Williams_{company}_{role}_{timestamp}.tex"
    out_file.write_text(tailored_tex, encoding="utf-8")
    update_tracker(company, role, out_file.name)

    cover_letter_file = generate_cover_letter(client, company, role, job_description, timestamp)

    print(f"Saved:  {out_file}")
    print(f"Saved:  {cover_letter_file}")
    print(f"Tracker updated: {TRACKER}")


if __name__ == "__main__":
    main()
